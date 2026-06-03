"""
Extract recipes from Meal Planner Excel workbooks and write .txt recipe files
into meal_planner/recipes/ for the meal planner app.

Usage:
  python scripts/populate_recipes_from_excel.py [path_4week.xlsx] [path_v2.xlsm]

If paths are omitted, uses the default paths below.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

# Default paths (edit if needed)
DEFAULT_4WEEK = Path(r"c:\Users\robev\Dropbox\Documents - Food\Meal Planner 4 week.xlsx")
DEFAULT_V2 = Path(r"c:\Users\robev\Dropbox\Documents - Food\Meal Planner v2.0.xlsm")

# Output directory relative to repo root
REPO_ROOT = Path(__file__).resolve().parent.parent
RECIPES_DIR = REPO_ROOT / "meal_planner" / "recipes"


def safe_basename(name: str) -> str:
    """Turn a meal name into a safe filename base (no path chars, no .txt)."""
    s = re.sub(r'[<>:"/\\|?*]', "", name)
    s = s.strip() or "recipe"
    s = "_".join(s.split())
    return s[:120] if len(s) > 120 else s


def write_recipe_txt(path: Path, name: str, ingredients: list[str], url: str | None = None) -> None:
    """Write a single recipe .txt in the format the app expects."""
    lines = [name]
    if url:
        lines.append(f"URL: {url}")
    if ingredients:
        lines.append("")
        for ing in ingredients:
            lines.append(f"- {ing}")
    path.write_text("\n".join(lines), encoding="utf-8")


def extract_from_v2(wb_path: Path) -> dict[str, list[str]]:
    """Read 'Meals n Ingredients' sheet: meal name -> list of ingredient lines (Item or Item (Amount))."""
    import openpyxl

    recipes: dict[str, list[str]] = {}
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    if "Meals n Ingredients" not in wb.sheetnames:
        wb.close()
        return recipes

    ws = wb["Meals n Ingredients"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        meal = (row[0] or "").strip()
        item = (row[1] or "").strip() if len(row) > 1 else ""
        amount = row[2] if len(row) > 2 and row[2] is not None else None
        if not meal:
            continue
        if meal not in recipes:
            recipes[meal] = []
        if item:
            if amount is not None and amount != "":
                recipes[meal].append(f"{item} ({amount})")
            else:
                recipes[meal].append(item)
    wb.close()
    return recipes


def extract_from_4week(wb_path: Path) -> dict[str, list[str]]:
    """Read Sheet1: column C = meal name, D = newline-separated ingredients."""
    import openpyxl

    recipes: dict[str, list[str]] = {}
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        meal = (row[2] or "").strip() if row[2] else ""
        raw_ing = row[3] if len(row) > 3 else None
        if not meal or meal.lower() == "take away":
            continue
        if meal in recipes:
            continue  # keep first occurrence
        ingredients = []
        if raw_ing and isinstance(raw_ing, str):
            for line in raw_ing.replace("\r", "\n").split("\n"):
                part = line.strip()
                if part:
                    ingredients.append(part)
        recipes[meal] = ingredients
    wb.close()
    return recipes


def main() -> None:
    if len(sys.argv) >= 3:
        path_4week = Path(sys.argv[1])
        path_v2 = Path(sys.argv[2])
    else:
        path_4week = DEFAULT_4WEEK
        path_v2 = DEFAULT_V2

    RECIPES_DIR.mkdir(parents=True, exist_ok=True)

    # Prefer v2.0 (richer ingredients); fill in any from 4 week that we don't have
    from_v2 = extract_from_v2(path_v2) if path_v2.exists() else {}
    from_4week = extract_from_4week(path_4week) if path_4week.exists() else {}

    seen_names: set[str] = set()
    used_basenames: set[str] = set()
    count = 0

    def unique_path(base: str) -> Path:
        candidate = base + ".txt"
        if candidate not in used_basenames:
            used_basenames.add(candidate)
            return RECIPES_DIR / candidate
        i = 2
        while True:
            candidate = f"{base}_{i}.txt"
            if candidate not in used_basenames:
                used_basenames.add(candidate)
                return RECIPES_DIR / candidate
            i += 1

    for name, ingredients in from_v2.items():
        base = safe_basename(name)
        target = unique_path(base)
        write_recipe_txt(target, name, ingredients)
        seen_names.add(name.strip().lower())
        count += 1

    for name, ingredients in from_4week.items():
        if name.strip().lower() in seen_names:
            continue
        base = safe_basename(name)
        target = unique_path(base)
        write_recipe_txt(target, name, ingredients)
        seen_names.add(name.strip().lower())
        count += 1

    print(f"Wrote {count} recipe files to {RECIPES_DIR}")


if __name__ == "__main__":
    main()
